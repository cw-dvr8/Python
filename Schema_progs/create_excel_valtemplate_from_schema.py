#!/usr/bin/env python3

"""
Program: create_excel_template_from_schema.py

Purpose: Use a JSON validation schema to generate an Excel workbook with
         the following structure:
         1) template
         2) dictionary
         3) possible column values

Input parameters: Full pathname to the JSON validation schema
                  Full pathname to the output template file

Outputs: Excel template workbook

Execution: create_excel_template_from_schema.py <JSON schema> <output file>

"""

import argparse
import collections
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from schema_tools import convert_bool_to_string, load_and_deref, values_list_keywords


def populate_dropdowns(col_location, col_key, values_sheet, values_column, values_dict, dropdown_dict):

    from openpyxl.utils.cell import get_column_letter, quote_sheetname

    column_range = ""
    max_row_count = "1048576"

    # Get the column letter to use when assigning the validation.
    column_letter = get_column_letter(col_location)
    column_range = column_letter + "2:" + column_letter + max_row_count

    # Create the validator and append it to the validator dictionary.
    values_column_letter = get_column_letter(values_column)
    values_start = f"${values_column_letter}${values_dict[col_key]['first_row']}"
    values_end = f"${values_column_letter}${values_dict[col_key]['last_row']}"
    formula_list = f"={values_sheet}!{values_start}:{values_end}"
    dropdown_dict[column_range] = DataValidation(type="list", formula1=formula_list, allow_blank=False,
                                                 showDropDown=True)

    return(dropdown_dict)


def main():

    parser = argparse.ArgumentParser()
    parser.add_argument("json_schema_file", type=argparse.FileType("r"),
                        help="Full pathname for the JSON schema file")
    parser.add_argument("output_file", type=str,
                        help="Full pathname for the output Excel file")

    args = parser.parse_args()

    values_list_keys = values_list_keywords()

    template_workbook = Workbook()
    header_column_dict = {}
    validation_dict = {}
    values_row_dict = collections.defaultdict(dict)

    # Load the JSON validation schema.
    _, json_schema = load_and_deref(args.json_schema_file)

    # Get the schema keys into a list and then write them to the template worksheet. The 
    # template worksheet will only have one row, but will have multiple columns.
    template_ws = template_workbook.active
    template_ws.title = "Template"

    column_number = 1
    for header_value in json_schema["properties"].keys():
        template_ws.cell(1, column_number).value = header_value

        # The header column dictionary is used to determine which column to set
        # validations on further down in the program.
        header_column_dict[header_value] = column_number
        column_number += 1

    # Write the dictionary to the dictionary worksheet. For keys that have a values list or
    # a regex, write the possible values to the values worksheet.
    dictionary_ws = template_workbook.create_sheet()
    dictionary_ws.title = "Dictionary"
    dictionary_ws.cell(1, 1).value = "key"
    dictionary_ws.cell(1, 2).value = "description"

    values_ws = template_workbook.create_sheet()
    values_ws.title = "Values"
    values_ws.cell(1, 1).value = "key"
    values_ws.cell(1, 2).value = "description"
    values_ws.cell(1, 3).value = "valueDescription"
    values_ws.cell(1, 4).value = "source"

    dictionary_row_number = 2
    values_row_number = 2

    for json_key in json_schema["properties"]:

        # Dictionary worksheet
        dictionary_ws.cell(dictionary_row_number, 1).value = json_key
        if "description" in json_schema["properties"][json_key]:
            dictionary_ws.cell(dictionary_row_number, 2).value = json_schema["properties"][json_key]["description"]
        dictionary_row_number += 1

        # Values worksheet
        if "pattern" in json_schema["properties"][json_key]:
            values_ws.cell(values_row_number, 1).value = json_key
            values_ws.cell(values_row_number, 2).value = json_schema["properties"][json_key]["pattern"]
            values_row_number += 1

        elif any([value_key in json_schema["properties"][json_key] for value_key in values_list_keys]):
            vkey = list(set(values_list_keys).intersection(json_schema["properties"][json_key]))[0]
            for vlist_row in json_schema["properties"][json_key][vkey]:
                if "const" in vlist_row:
                    if json_key not in values_row_dict:
                        values_row_dict[json_key]["first_row"] = values_row_number

                    values_row_dict[json_key]["last_row"] = values_row_number

                    values_ws.cell(values_row_number, 1).value = json_key

                    # If the value is a Boolean, we have to convert it to a string; otherwise
                    # Excel will force it into all-caps, i.e. (TRUE, FALSE) and this is not
                    # what we want. The function converts Booleans, but if the value is not a
                    # Boolean it will return the original value.
                    values_ws.cell(values_row_number, 2).value = convert_bool_to_string(vlist_row["const"])

                    if "description" in vlist_row:
                        values_ws.cell(values_row_number, 3).value = vlist_row["description"]
                    if "source" in vlist_row:
                        values_ws.cell(values_row_number, 4).value = vlist_row["source"]

                    values_row_number += 1

    # Set the column validations.
    if len(values_row_dict) > 0:
        for val_column in values_row_dict:
            validation_dict = populate_dropdowns(header_column_dict[val_column], val_column, "Values", 2,
                                                 values_row_dict, validation_dict)

    if len(validation_dict) > 0:
        for col_range in validation_dict:
            validation_dict[col_range].ranges.add(col_range)
            template_ws.add_data_validation(validation_dict[col_range])

    template_workbook.save(args.output_file)


if __name__ == "__main__":
    main()
