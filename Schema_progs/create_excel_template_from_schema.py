#!/usr/bin/env python3

"""
Program: create_excel_template_from_schema.py

Purpose: Use a JSON validation schema to generate an Excel workbook with
         the following structure:
         1) template
         2) dictionary
         3) possible column values

Input parameters: Full pathname to the JSON validation schema
                  Full pathname to the output template file

Outputs: Excel template workbook

Execution: create_excel_template_from_schema.py <JSON schema> <output file>

"""

import argparse
from openpyxl import Workbook
from schema_tools import convert_bool_to_string, get_schema_properties

def main():

    parser = argparse.ArgumentParser()
    parser.add_argument("json_schema_file", type=argparse.FileType("r"),
                        help="Full pathname for the JSON schema file")
    parser.add_argument("output_file", type=str,
                        help="Full pathname for the output Excel file")

    args = parser.parse_args()

    template_workbook = Workbook()

    definitions, values = get_schema_properties(args.json_schema_file)

    # Get the schema keys into a list and then write them to the template worksheet. The 
    # template worksheet will only have one row, but will have multiple columns.
    template_ws = template_workbook.active
    template_ws.title = "Template"

    column_number = 1
    for header_value in definitions.keys():
        template_ws.cell(1, column_number).value = header_value
        column_number += 1

    # Write the dictionary to the dictionary worksheet. For keys that have a values list or
    # a regex, write the possible values to the values worksheet.
    dictionary_ws = template_workbook.create_sheet()
    dictionary_ws.title = "Dictionary"
    dictionary_ws.cell(1, 1).value = "key"
    dictionary_ws.cell(1, 2).value = "description"

    values_ws = template_workbook.create_sheet()
    values_ws.title = "Values"
    values_ws.cell(1, 1).value = "key"
    values_ws.cell(1, 2).value = "description"
    values_ws.cell(1, 3).value = "valueDescription"
    values_ws.cell(1, 4).value = "source"

    dictionary_row_number = 2
    values_row_number = 2

    for schema_key in definitions:

        # Dictionary worksheet
        dictionary_ws.cell(dictionary_row_number, 1).value = schema_key
        dictionary_ws.cell(dictionary_row_number, 2).value = definitions[schema_key]["description"]
        dictionary_row_number += 1

        # Values worksheet
        if schema_key in values:
            for vlist_row in values[schema_key]:
                values_ws.cell(values_row_number, 1).value = schema_key

                # If the value is a Boolean, we have to convert it to a string; otherwise
                # Excel will force it into all-caps, i.e. (TRUE, FALSE) and this is not
                # what we want. The function converts Booleans, but if the value is not a
                # Boolean it will return the original value.
                values_ws.cell(values_row_number, 2).value = convert_bool_to_string(vlist_row["value"])
                values_ws.cell(values_row_number, 3).value = vlist_row["valueDescription"]
                values_ws.cell(values_row_number, 4).value = vlist_row["source"]
                values_row_number += 1

    template_workbook.save(args.output_file)


if __name__ == "__main__":
    main()
